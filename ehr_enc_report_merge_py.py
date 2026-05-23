from openpyxl import load_workbook
import pandas as pd
from pathlib import Path

# ==============================
# CONFIG
# ==============================
INPUT_FILE = "comprehensive_professional_and_other_ehr_fullmarch.xlsx"
OUTPUT_FILE = "comprehensive_professional_and_other_ehr_fullmarch_merged.xlsx"

FINAL_COLUMNS = [
    "Site(s) of Service",
    "Patient Name",
    "Method of Interaction",
    "Payer Plan",
    "Service Date",
    "Primary Medical Dx",
    "Other Medical Dx Codes",
    "Primary Treatment Dx",
    "Other Treatment Dx Codes",
    "Discipline",
    "CPT Code",
    "CPT Description",
    "Total Units",
    "Total Minutes",
    "Total Days",
]

# ==============================
# LOAD WORKBOOK
# ==============================
wb = load_workbook(INPUT_FILE, data_only=True)

all_rows = []

# ==============================
# PROCESS EACH SHEET
# ==============================
for ws in wb.worksheets:

    current_site = None

    for row in ws.iter_rows(values_only=True):

        if not row:
            continue

        # ----------------------------------
        # Detect Site of Service header
        # ----------------------------------
        if isinstance(row[0], str) and row[0].strip() == "Site of Service:":
            current_site = row[4]  # Column E
            continue

        # ----------------------------------
        # Extract ONLY patient rows
        # ----------------------------------
        if not isinstance(row[0], str) or "," not in row[0]:
            continue

        record = {
            "Site(s) of Service": current_site,
            "Patient Name": row[0],
            "Method of Interaction": row[4],
            "Payer Plan": row[6],
            "Service Date": row[9],
            "Primary Medical Dx": row[11],
            "Other Medical Dx Codes": row[13],
            "Primary Treatment Dx": row[16],
            "Other Treatment Dx Codes": row[18],
            "Discipline": row[21],
            "CPT Code": row[23],
            "CPT Description": row[25],
            "Total Units": row[27],
            "Total Minutes": row[29],
            "Total Days": row[31],
        }

        all_rows.append(record)

# ==============================
# SAFETY CHECK
# ==============================
if not all_rows:
    raise RuntimeError("❌ No patient rows extracted — check Excel structure.")

# ==============================
# CREATE DATAFRAME
# ==============================
df = pd.DataFrame(all_rows)
df = df[FINAL_COLUMNS]

# ==============================
# WRITE OUTPUT
# ==============================
df.to_excel(OUTPUT_FILE, index=False)

print(f"✅ Rows written: {len(df)}")
print(f"📄 Output file: {Path(OUTPUT_FILE).resolve()}")
