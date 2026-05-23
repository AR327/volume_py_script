from openpyxl import load_workbook
import pandas as pd

# ===== INPUT / OUTPUT FILE PATHS =====
INPUT_FILE = "153341_appt_ehr_report_may22_25.xlsx"

OUTPUT_FILE = "153341_appt_ehr_report_may22_25_merged.xlsx"

# ====================================
wb = load_workbook(INPUT_FILE, read_only=True, data_only=True)

rows = []

for ws in wb.worksheets:
    # Sheet-level values
    patient_name = ws["A2"].value    # Row 2
    facility = ws["A6"].value        # Row 6

    # Headers from row 12
    headers = [cell.value for cell in ws[12]]

    # Data rows from row 13 onward
    for r in ws.iter_rows(min_row=13, values_only=True):
        if all(v is None for v in r):
            continue

        row = dict(zip(headers, r))
        row["Patient Name"] = patient_name
        row["Facility"] = facility
        rows.append(row)

# Create consolidated DataFrame
df = pd.DataFrame(rows)

# Export
df.to_excel(OUTPUT_FILE, index=False)

print(f"✅ Consolidated file created: {OUTPUT_FILE}")
